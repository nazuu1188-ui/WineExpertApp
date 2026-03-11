import openpyxl
import pandas as pd


def _read_rect_table(ws, header_row: int, start_col: int, max_cols: int = 200) -> pd.DataFrame:
    """
    Reads a rectangular table:
    - headers on `header_row` starting at `start_col`
    - data rows continue until a fully blank row
    """
    # headers
    headers = []
    c = start_col
    while c <= max_cols:
        v = ws.cell(header_row, c).value
        if v is None or str(v).strip() == "":
            if headers:
                break
            c += 1
            continue
        headers.append(str(v).strip())
        c += 1

    # rows
    data = []
    r = header_row + 1
    while r <= ws.max_row:
        row = [ws.cell(r, start_col + i).value for i in range(len(headers))]
        if all(v is None or str(v).strip() == "" for v in row):
            break
        data.append(row)
        r += 1

    return pd.DataFrame(data, columns=headers)


def load_redwine_oam_workbook(xlsx_file) -> dict[str, pd.DataFrame]:
    """
    Tailored to your 'RedWineQuality OAM.xlsx' layout:

    Expected sheets:
      - Raw Data
      - Attributes
      - OAM
      - models
      - Conclusion

    Extracts:
      - raw_data (pandas read_excel)
      - attributes (pandas read_excel)
      - conclusion table (row 4, col B)
      - models table (row 7, col A)
      - oam_matrix (SM rows from OAM sheet: row 12+, col B..O)
    """
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)

    out: dict[str, pd.DataFrame] = {}

    # Clean tabular sheets
    out["raw_data"] = pd.read_excel(xlsx_file, sheet_name="Raw Data")
    out["attributes"] = pd.read_excel(xlsx_file, sheet_name="Attributes")

    # Conclusion table (header at row 4, starts at column B)
    out["conclusion"] = _read_rect_table(wb["Conclusion"], header_row=4, start_col=2)

    # models table (header at row 7, starts at column A)
    out["models"] = _read_rect_table(wb["models"], header_row=7, start_col=1)

    # OAM matrix (SM rows)
    ws_oam = wb["OAM"]

    # Attribute IDs row in your file: row 9, columns C..N = A1..A12
    attr_ids = []
    for col in range(3, 15):  # C=3 .. N=14
        v = ws_oam.cell(9, col).value
        attr_ids.append(str(v).strip() if v is not None else f"Col{col}")

    # Data rows start at row 12:
    # B = SMxxx, C..N = values, O = DuplicateSignature
    rows = []
    r = 12
    while r <= ws_oam.max_row:
        oam_id = ws_oam.cell(r, 2).value  # column B
        if oam_id is None or str(oam_id).strip() == "":
            break

        values = [ws_oam.cell(r, c).value for c in range(3, 15)]  # C..N
        dup_sig = ws_oam.cell(r, 15).value  # O
        rows.append([str(oam_id).strip()] + values + [dup_sig])
        r += 1

    oam_cols = ["OAM"] + attr_ids + ["DuplicateSignature"]
    out["oam_matrix"] = pd.DataFrame(rows, columns=oam_cols)

    return out